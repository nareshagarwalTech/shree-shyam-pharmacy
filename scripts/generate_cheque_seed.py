"""
Generate a one-shot SQL seed file from SHREE_SHYAM_CHEQUE (1).xlsx.

Reads the Excel workbook, normalises the rows into parties + cheques, and
writes scripts/seed_cheques.sql which can be run in the Supabase SQL editor
AFTER migration 010 has been applied.

Usage:
    python scripts/generate_cheque_seed.py "C:\\path\\to\\SHREE_SHYAM_CHEQUE (1).xlsx"

If no path is given, defaults to the OneDrive path the user shared.
"""
from __future__ import annotations

import sys
import datetime
from pathlib import Path

import openpyxl

DEFAULT_INPUT = r"C:\Users\nares\OneDrive\Documents\Shyam\SHREE_SHYAM_CHEQUE (1).xlsx"
DEFAULT_OUTPUT = Path(__file__).parent / "seed_cheques.sql"

# Status mapping (Excel -> our enum)
STATUS_MAP = {
    "CLEARED":   "cleared",
    "CANCELLED": "cancelled",
    "BOUNCED":   "bounced",
    "PENDING":   "pending",
}

# Party names that should NOT become real party rows. In the Excel, "CANCEL"
# was used as a placeholder when a cheque from the cheque-book was voided
# without ever being given to anyone. We translate those into orphan
# cheques (party_id NULL) with status='cancelled' instead.
PARTY_BLACKLIST = {"CANCEL", "SELF CHANGE"}

# Crude category guess based on the name. Override later in /parties.
def guess_category(name: str) -> str:
    n = name.upper()
    if any(k in n for k in (" PHARMA", "DRUGS", "AGENCIES", "DISTRIBUTOR", "MEDICAL", "ENTERPRISES",
                            "ASSOCIATES", "TRADERS", "MARKETING", "HEALTH", "PITTI", "SURGICAL",
                            "DETTOL", "GODREJ", "J&J", "LIBERTY", "VOLINI", "HORLICKS")):
        return "pharma"
    if name.startswith("B.") or name.startswith("MR.") or name.startswith("MS."):
        return "staff"
    if any(k in n for k in ("LOGISTICS", "MSWIPE", "PAY ONE", "RETAIL")):
        return "service"
    return "other"


def sql_str(s: str | None) -> str:
    if s is None:
        return "NULL"
    return "'" + s.replace("'", "''") + "'"


def sql_date(d) -> str:
    if d is None:
        return "NULL"
    if isinstance(d, datetime.datetime):
        d = d.date()
    if isinstance(d, datetime.date):
        return f"'{d.isoformat()}'"
    # try parse
    s = str(d).strip()
    return sql_str(s)


def sql_num(n) -> str:
    if n is None or n == "":
        return "NULL"
    return f"{float(n):.2f}"


def main() -> int:
    args = sys.argv[1:]
    inp = Path(args[0]) if args else Path(DEFAULT_INPUT)
    if not inp.exists():
        print(f"ERROR: input file not found: {inp}", file=sys.stderr)
        return 1

    out = DEFAULT_OUTPUT
    print(f"Reading: {inp}")
    wb = openpyxl.load_workbook(str(inp), data_only=True)

    # ---- 1. Parties from PartyList sheet ----
    ws = wb["PartyList"]
    parties: list[str] = []
    seen: set[str] = set()
    for r in range(3, ws.max_row + 1):
        n = ws.cell(row=r, column=1).value
        if not n:
            continue
        name = str(n).strip()
        if not name or name.upper() in PARTY_BLACKLIST:
            continue
        if name not in seen:
            seen.add(name)
            parties.append(name)

    print(f"Parties to seed: {len(parties)}")

    # ---- 2. Cheques from Daily_Cheque sheet ----
    ws = wb["Daily_Cheque"]
    cheques: list[dict] = []
    last = 0
    for r in range(3, ws.max_row + 1):
        if ws.cell(row=r, column=2).value or ws.cell(row=r, column=3).value:
            last = r

    for r in range(3, last + 1):
        date_v   = ws.cell(row=r, column=1).value
        party_v  = ws.cell(row=r, column=2).value
        cheq_v   = ws.cell(row=r, column=3).value
        amt_v    = ws.cell(row=r, column=4).value
        bank_v   = ws.cell(row=r, column=5).value
        status_v = ws.cell(row=r, column=6).value
        clear_v  = ws.cell(row=r, column=7).value
        remarks_v = ws.cell(row=r, column=9).value

        # Skip header / blank rows
        if not party_v and not cheq_v and not amt_v:
            continue

        party = str(party_v).strip() if party_v else None
        is_orphan = party and party.upper() in PARTY_BLACKLIST

        if amt_v in (None, "", 0):
            # cancelled cheque-book entries can have no amount
            amt = None
        else:
            try:
                amt = float(amt_v)
            except Exception:
                amt = None

        cheq_no_str: str | None = None
        is_online = False
        if cheq_v is not None:
            s = str(cheq_v).strip()
            if s.upper() == "ONLINE":
                is_online = True
            else:
                # numeric; strip ".0" suffix from openpyxl
                try:
                    cheq_no_str = str(int(float(s)))
                except Exception:
                    cheq_no_str = s or None

        st = STATUS_MAP.get(str(status_v).strip().upper(), "pending") if status_v else "pending"
        if is_orphan:
            st = "cancelled"

        cheques.append({
            "issue_date":     date_v,
            "party":          None if is_orphan else party,
            "is_online":      is_online,
            "cheque_no":      cheq_no_str,
            "amount":         amt,
            "status":         st,
            "clearance_date": clear_v if isinstance(clear_v, (datetime.date, datetime.datetime)) else None,
            "remarks":        str(remarks_v).strip() if remarks_v else None,
            "bank":           str(bank_v).strip() if bank_v else None,
        })

    # Cheques without an amount or issue_date can't satisfy NOT NULL constraints —
    # drop them with a warning.
    valid = [c for c in cheques if c["issue_date"] and c["amount"] and c["amount"] > 0]
    skipped = len(cheques) - len(valid)
    print(f"Cheques: {len(valid)} valid, {skipped} skipped (missing date/amount)")

    # ---- 3. Write SQL ----
    sql: list[str] = []
    sql.append("-- =============================================================================")
    sql.append("-- One-shot seed: parties + cheques from SHREE_SHYAM_CHEQUE (1).xlsx")
    sql.append("-- Generated by scripts/generate_cheque_seed.py")
    sql.append("-- Pre-req: migration 010_cheques.sql must be applied first.")
    sql.append("-- Idempotent: ON CONFLICT DO NOTHING on parties; safe to re-run for parties,")
    sql.append("-- but cheques will duplicate so wipe first if re-running:")
    sql.append("--     DELETE FROM cheques;")
    sql.append("-- =============================================================================")
    sql.append("")
    sql.append("BEGIN;")
    sql.append("")
    sql.append("-- 1. Parties")
    for p in parties:
        cat = guess_category(p)
        sql.append(
            f"INSERT INTO parties (name, category) VALUES "
            f"({sql_str(p)}, {sql_str(cat)}) ON CONFLICT (name) DO NOTHING;"
        )

    sql.append("")
    sql.append("-- 2. Cheques (looked up by party name; bank_id falls back to default bank)")
    sql.append("DO $$")
    sql.append("DECLARE")
    sql.append("  v_default_bank UUID := (SELECT id FROM banks WHERE is_default = true LIMIT 1);")
    sql.append("  v_party UUID;")
    sql.append("BEGIN")

    for c in valid:
        party_lookup = (
            f"(SELECT id FROM parties WHERE name = {sql_str(c['party'])} LIMIT 1)"
            if c["party"] else "NULL"
        )
        sql.append(
            "  INSERT INTO cheques (party_id, bank_id, is_online, cheque_no, amount, "
            "issue_date, clearance_date, status, remarks) VALUES ("
            f"{party_lookup}, "
            "v_default_bank, "
            f"{'true' if c['is_online'] else 'false'}, "
            f"{sql_str(c['cheque_no'])}, "
            f"{sql_num(c['amount'])}, "
            f"{sql_date(c['issue_date'])}, "
            f"{sql_date(c['clearance_date'])}, "
            f"{sql_str(c['status'])}, "
            f"{sql_str(c['remarks'])}"
            ");"
        )

    sql.append("END $$;")
    sql.append("")
    sql.append("COMMIT;")
    sql.append("")
    sql.append("-- Verification")
    sql.append("SELECT 'parties' AS t, COUNT(*) FROM parties")
    sql.append("UNION ALL SELECT 'cheques', COUNT(*) FROM cheques")
    sql.append("UNION ALL SELECT 'cheques pending',  COUNT(*) FROM cheques WHERE status = 'pending'")
    sql.append("UNION ALL SELECT 'cheques cleared',  COUNT(*) FROM cheques WHERE status = 'cleared'")
    sql.append("UNION ALL SELECT 'cheques cancelled', COUNT(*) FROM cheques WHERE status = 'cancelled';")
    sql.append("")

    out.write_text("\n".join(sql), encoding="utf-8")
    print(f"Wrote: {out}")
    print(f"Parties: {len(parties)}    Cheques: {len(valid)}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
